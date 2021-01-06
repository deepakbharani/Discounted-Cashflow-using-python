# Author:   Bharani Deepak
# Date:     03.01.2021
# Title:    Stock Valuation using Discounted Cashflow method 

import openpyxl
import numpy as np
import requests
import bs4

def dcf(wb,stock_price, beta, shares_outstanding):

    IncomeStmt_mod   = wb['IncomeStatement']                            # Load income statement of modified data
    balancesheet_mod = wb['BalanceSheet']                               # Load Balance sheet of modified data
    Cashflowstmt_mod = wb['CashFlow']                                   # Load Cash flow statement of modified data
    num_period = 4
    years_to_forecast = 5
    risk_free_return = 4
    expected_return = 10
    corp_tax = 30                                                       # Corporate tax
    discount_factor = []
    present_value = []
    perpetual_growth = 2.5

    for i in range (1,len(balancesheet_mod['A'])):

        # Read long term debt
        if balancesheet_mod.cell(row = i, column = 1).value == 'Long-term debt':
            long_term_debt = balancesheet_mod.cell(row = i, column = 2).value

        if balancesheet_mod.cell(row = i, column = 1).value == 'Total stockholders\' equity':
            equity_capital = balancesheet_mod.cell(row = i, column = 2).value

    for i in range(1, len(IncomeStmt_mod['A'])):

        if IncomeStmt_mod.cell(row = i, column = 1).value == 'Total revenue':
            total_revenue = np.array([IncomeStmt_mod.cell(row = i, column = 6).value,
                                      IncomeStmt_mod.cell(row = i, column = 5).value,
                                      IncomeStmt_mod.cell(row = i, column = 4).value,
                                      IncomeStmt_mod.cell(row = i, column = 3).value])

        if IncomeStmt_mod.cell(row = i, column = 1).value == 'Net income':
            net_income = np.array([IncomeStmt_mod.cell(row = i, column = 6).value,
                                   IncomeStmt_mod.cell(row = i, column = 5).value,
                                   IncomeStmt_mod.cell(row = i, column = 4).value,
                                   IncomeStmt_mod.cell(row = i, column = 3).value])
        # Read interest expense
        if IncomeStmt_mod.cell(row = i, column = 1).value == 'Interest expense':
            interest_expenses = IncomeStmt_mod.cell(row = i, column = 2).value

    free_cash_flow = np.array([Cashflowstmt_mod.cell(row=len(Cashflowstmt_mod['A']), column=6).value,
                               Cashflowstmt_mod.cell(row=len(Cashflowstmt_mod['A']), column=5).value,
                               Cashflowstmt_mod.cell(row=len(Cashflowstmt_mod['A']), column=4).value,
                               Cashflowstmt_mod.cell(row=len(Cashflowstmt_mod['A']), column=3).value])

    # Calculate CAGR of Total Revenue-------------------------------------------------------------------->
    # Total revenue growth rate -> tot_rev_cagr
    total_revenue_growth_rate = []

    total_revenue_growth_rate.append(((total_revenue[1]-total_revenue[0])/total_revenue[0])*100)
    total_revenue_growth_rate.append(((total_revenue[2]-total_revenue[1])/total_revenue[1])*100)
    total_revenue_growth_rate.append(((total_revenue[3]-total_revenue[2])/total_revenue[2])*100)

    tot_rev_cagr = sum(total_revenue_growth_rate)/len(total_revenue_growth_rate)
    # Calculate ratio of net income to total revenue----------------------------------------------------->
    net_income_to_total_revenue = np.divide(net_income,total_revenue) * 100
    avg_net_income_to_total_revenue = np.mean(net_income_to_total_revenue)

    # Calculate ratio of free cash flow to net income---------------------------------------------------->
    free_cash_flow_to_net_income = np.divide(free_cash_flow,net_income) * 100
    avg_free_cash_flow_to_net_income = np.mean(free_cash_flow_to_net_income)

    for i in range (years_to_forecast):

        # forecast total revenue for next few years (in this case next 5 years)
        revenue_forecast = total_revenue[-1] + ((total_revenue[-1]*tot_rev_cagr) / 100)
        total_revenue = np.append(total_revenue,revenue_forecast)

        # forecast net income for next few years (in this case next 5 years)
        net_income_forecast = net_income[-1] + ((net_income[-1] * avg_net_income_to_total_revenue) / 100)
        net_income = np.append(net_income, net_income_forecast)

        # forecast free cash flow for next few years (in this case next 5 years)
        free_cash_flow_forecast = np.append(free_cash_flow,net_income[-1] * avg_free_cash_flow_to_net_income / 100)
        free_cash_flow = np.append(free_cash_flow, free_cash_flow_forecast)

        # Calculate Discount factor
        discount_factor.append((1+(expected_return/100))**(i+1))

        # Calculate present value of future free cash flow
        present_value.append(free_cash_flow[-1] / discount_factor[i])

    rate_of_interest_expense = (interest_expenses / long_term_debt) * 100

    # Calculate WACC
    total_equity = equity_capital + long_term_debt
    weight_equity_capital = equity_capital / total_equity
    weight_debt_capital = long_term_debt / total_equity
    ror_equity_capital = risk_free_return + (beta*(expected_return-risk_free_return))       # ROR = Rate of return
    ror_debt_capital = rate_of_interest_expense*(1-(corp_tax/100))
    wacc = (weight_equity_capital*ror_equity_capital) + (weight_debt_capital*ror_debt_capital)

    # Calculate terminal value of free cash flow
    terminal_value = free_cash_flow[-1]*(1+(perpetual_growth/100)) / ((wacc/100)-(perpetual_growth/100))
    present_terminal_value = terminal_value/discount_factor[-1]
    today_value = (sum(present_value) + present_terminal_value)*0.0001

    print('Total Revenuee:',total_revenue)
    print('Net income:',net_income)
    print('Discount Factor:',discount_factor)
    print('Present Value:',present_value)
    print('Today\'s Value:',today_value)
    print('Total revenue growth rate:',tot_rev_cagr)
    print('Net income to total revenue:',avg_net_income_to_total_revenue)
    print('Free cash flow to net income:',avg_free_cash_flow_to_net_income)
    print('WACC:', wacc)
    print('ror equity:',ror_equity_capital)
    print('ror debt capital:',ror_debt_capital)
    print('weight equity:',weight_equity_capital,'weight debt:',weight_debt_capital)
    print('Interest expense:',interest_expenses)
    print('Long term debt:',long_term_debt)
    print('Equity capital:',equity_capital)
    print('Interest expense:',rate_of_interest_expense)

    # Calculate intrinsic value
    intrinsic_value = today_value / shares_outstanding

    # Calculate margin of safety
    margin_of_safty = ((intrinsic_value-stock_price)/stock_price)*100

    return intrinsic_value, margin_of_safty

def main():

    wb_mod = openpyxl.load_workbook('Kaveri_modified.xlsx')             # Open formated excel
    stock_name = 'Kaveri Seeds'
    stock_price = 546
    beta = 0.6
    shares_outstanding = 6.01
    [intrinsic_value, margin_of_safty] = dcf(wb_mod,stock_price, beta, shares_outstanding)

    print('The intrinsic value is:',intrinsic_value,'and the margin of safety is:',margin_of_safty,'%')

if __name__ == '__main__':
    main()